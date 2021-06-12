import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

'''
load in the name of excel file------------------------------------------
'''
wb = load_workbook('excelFileName.xlsx')

'''
write the name of sheet to be converted to json 
Sheet1, Sheet2, etc are default-----------------------------------------
'''
sheet = wb['Sheet1']

'''
init list to hold dict--------------------------------------------------
'''

listToHoldDict=[]


'''
loop over each row and fetch values into Dict
'''
for row in islice(sheet.values, 1, sheet.max_row):
    Dict = OrderedDict()
    ''' name of rows add in '''
    Dict['nameOfRow1'] = row[0]
    Dict['nameOfRow2'] = row[1]
    Dict['nameOfRow3'] = row[2]

    listToHoldDict.append(Dict)

j = json.dumps(listToHoldDict)

with open('data.json', 'w') as f:
    f.write(j)